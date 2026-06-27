#!/usr/bin/env python3
"""
migrate.py
──────────────────────────────────────────────────────────────────────────────
One-time migration: contacts.xlsx → Google Sheets

Creates three sheet tabs:
  - Contacts   (cleaned 139-contact dataset)
  - Users      (with default admin account)
  - AuditLog   (headers only)

Usage:
  # Via environment variables (preferred):
  GOOGLE_SHEETS_SPREADSHEET_ID=xxx \
  GOOGLE_SHEETS_CREDENTIALS_JSON='{"type":"service_account",...}' \
  ADMIN_EMAIL=admin@assocham.org \
  ADMIN_PASSWORD=ChangeMe123 \
  python migrate.py contacts.xlsx

  # Via CLI arguments:
  python migrate.py contacts.xlsx \
    --spreadsheet-id xxx \
    --credentials /path/to/creds.json \
    --admin-email admin@assocham.org \
    --admin-password ChangeMe123
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import uuid
from datetime import date, datetime
from pathlib import Path


# ── Contacts sheet column order ───────────────────────────────────────────────

CONTACTS_HEADERS = [
    "Contact_ID", "Name", "Designation", "Company", "Parent_Organization",
    "Sector", "Company_Type", "Address", "City", "State", "Country",
    "Phone", "Alternate_Phone", "Email", "Alternate_Email",
    "Website", "LinkedIn", "Contact_Priority", "Status",
    "Event_Source", "Source_Type", "Import_Source",
    "Date_Added", "Last_Updated", "Next_Followup_Date",
    "AI_Tags", "AI_Summary", "Notes", "Raw_Extraction_JSON",
]

USERS_HEADERS = [
    "User_ID", "Name", "Email", "Password_Hash", "Role",
    "Department", "Status", "Created_At", "Last_Login",
]

AUDIT_HEADERS = [
    "Log_ID", "Timestamp", "User_ID", "User_Name",
    "Action", "Contact_ID", "Contact_Name", "Details",
]

# Map from xlsx column names → our schema column names
COLUMN_RENAMES = {
    "Contact ID": "Contact_ID",
    "Parent Organization": "Parent_Organization",
    "Company Type": "Company_Type",
    "Alternate Phone": "Alternate_Phone",
    "Alternate Email": "Alternate_Email",
    "Contact Priority": "Contact_Priority",
    "Notes/Services": "Notes",
    "Sector_Standardized": "Sector",   # prefer standardized over raw
    "Event_Source": "Event_Source",
}

# Columns to DROP from xlsx (formula columns, duplicates)
COLUMNS_TO_DROP = {"Sector", "Duplicate_Check"}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Migrate contacts.xlsx to Google Sheets")
    p.add_argument("xlsx_path", nargs="?", default="contacts.xlsx", help="Path to contacts.xlsx")
    p.add_argument("--spreadsheet-id", default=os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID", ""))
    p.add_argument("--credentials", default=os.getenv("GOOGLE_SHEETS_CREDENTIALS_JSON", ""))
    p.add_argument("--contacts-sheet", default=os.getenv("GOOGLE_SHEETS_CONTACTS_SHEET_NAME", "Contacts"))
    p.add_argument("--users-sheet", default=os.getenv("GOOGLE_SHEETS_USERS_SHEET_NAME", "Users"))
    p.add_argument("--audit-sheet", default=os.getenv("GOOGLE_SHEETS_AUDIT_SHEET_NAME", "AuditLog"))
    p.add_argument("--admin-email", default=os.getenv("ADMIN_EMAIL", "admin@assocham.org"))
    p.add_argument("--admin-password", default=os.getenv("ADMIN_PASSWORD", "ChangeMe123!"))
    p.add_argument("--admin-name", default=os.getenv("ADMIN_NAME", "ASSOCHAM Admin"))
    p.add_argument("--dry-run", action="store_true", help="Print rows without writing to Sheets")
    return p.parse_args()


def clean_phone(raw: str) -> str:
    """Strip Excel formula prefix =+ and normalize phone numbers."""
    if not raw:
        return ""
    s = str(raw).strip()
    # Remove Excel formula prefix
    s = re.sub(r"^=\s*\+?", "", s)
    s = re.sub(r"^=\s*", "", s)
    # Keep digits, +, spaces, hyphens
    s = re.sub(r"[^\d+\s\-\(\)]", "", s)
    s = s.strip()
    # Ensure +91 prefix for Indian numbers
    digits = re.sub(r"\D", "", s)
    if len(digits) == 10 and not s.startswith("+"):
        s = "+91" + digits
    elif len(digits) == 12 and digits.startswith("91") and not s.startswith("+"):
        s = "+" + digits
    return s


def clean_row(raw: dict, row_num: int) -> dict:
    """Clean and normalize a single contact row from the xlsx."""
    cleaned = {}

    # Rename and drop columns
    for k, v in raw.items():
        k_str = str(k).strip()
        if k_str in COLUMNS_TO_DROP:
            continue
        new_key = COLUMN_RENAMES.get(k_str, k_str.replace(" ", "_"))
        val = str(v).strip() if v is not None and str(v).strip() not in ("None", "nan", "") else ""
        cleaned[new_key] = val

    # Clean phones
    for phone_field in ("Phone", "Alternate_Phone"):
        if cleaned.get(phone_field):
            cleaned[phone_field] = clean_phone(cleaned[phone_field])

    # Generate Contact_ID if missing
    if not cleaned.get("Contact_ID"):
        cleaned["Contact_ID"] = f"CNT-{row_num:03d}"
    elif not cleaned["Contact_ID"].startswith("CNT-"):
        # Normalize existing IDs like C001 → CNT-001
        num_part = re.sub(r"\D", "", cleaned["Contact_ID"])
        cleaned["Contact_ID"] = f"CNT-{int(num_part):03d}" if num_part else f"CNT-{row_num:03d}"

    # Defaults
    if not cleaned.get("Status"):
        cleaned["Status"] = "Active"
    if not cleaned.get("Country"):
        cleaned["Country"] = "India"
    if not cleaned.get("Contact_Priority"):
        cleaned["Contact_Priority"] = "Medium"
    if not cleaned.get("Source_Type"):
        cleaned["Source_Type"] = "Excel Import"
    if not cleaned.get("Import_Source"):
        cleaned["Import_Source"] = "contacts.xlsx migration"

    today = date.today().isoformat()
    if not cleaned.get("Date_Added"):
        cleaned["Date_Added"] = today
    cleaned["Last_Updated"] = today

    # Ensure Raw_Extraction_JSON is empty for migrated contacts
    cleaned.setdefault("Raw_Extraction_JSON", "")

    return cleaned


def row_to_sheet_values(contact: dict) -> list[str]:
    return [str(contact.get(h, "") or "") for h in CONTACTS_HEADERS]


def build_sheets_service(credentials: str):
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build

    SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
    stripped = credentials.strip()
    if stripped.startswith("{"):
        creds_data = json.loads(stripped)
    else:
        path = Path(stripped).expanduser()
        if not path.is_file():
            sys.exit(f"❌ Credentials file not found: {path}")
        with open(path) as f:
            creds_data = json.load(f)
    creds = Credentials.from_service_account_info(creds_data, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def ensure_sheet_tab(service, spreadsheet_id: str, title: str):
    """Create sheet tab if it doesn't already exist."""
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing = [s["properties"]["title"] for s in meta.get("sheets", [])]
    if title not in existing:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": title}}}]},
        ).execute()
        print(f"  ✓ Created sheet tab: {title}")
    else:
        print(f"  ℹ Sheet tab already exists: {title}")


def write_to_sheet(service, spreadsheet_id: str, sheet_name: str, rows: list[list[str]]):
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'",
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!A1",
        valueInputOption="USER_ENTERED",
        body={"values": rows},
    ).execute()
    print(f"  ✓ Written {len(rows) - 1} data rows to '{sheet_name}'")


def create_admin_user(password: str, name: str, email: str) -> list[str]:
    from passlib.context import CryptContext
    pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
    hashed = pwd_ctx.hash(password)
    now = datetime.utcnow().isoformat()
    return ["USR-001", name, email, hashed, "Admin", "Management", "Active", now, ""]


def main():
    args = parse_args()

    if not args.spreadsheet_id and not args.dry_run:
        sys.exit("❌ --spreadsheet-id or GOOGLE_SHEETS_SPREADSHEET_ID is required.")
    if not args.credentials and not args.dry_run:
        sys.exit("❌ --credentials or GOOGLE_SHEETS_CREDENTIALS_JSON is required.")

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.is_file():
        sys.exit(f"❌ File not found: {xlsx_path}")

    print(f"\n{'='*60}")
    print(f"ASSOCHAM CRM — Migration Script")
    print(f"{'='*60}")
    print(f"Source file : {xlsx_path}")
    print(f"Spreadsheet : {args.spreadsheet_id or '(dry-run)'}")
    print(f"Contacts tab: {args.contacts_sheet}")
    print(f"Users tab   : {args.users_sheet}")
    print(f"Audit tab   : {args.audit_sheet}")
    print(f"Dry run     : {args.dry_run}")
    print()

    # ── Read xlsx ──────────────────────────────────────────────────────────────
    print("📖 Reading contacts.xlsx ...")
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("❌ openpyxl not installed. Run: pip install openpyxl")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter)]
    print(f"  Found {len(headers)} columns: {headers}")

    contacts_rows = [CONTACTS_HEADERS]  # Header row first
    seen_ids: set[str] = set()
    skipped = 0
    processed = 0

    for row_num, row in enumerate(rows_iter, start=2):
        raw = dict(zip(headers, row))
        # Skip completely empty rows
        if all(not str(v or "").strip() for v in raw.values()):
            skipped += 1
            continue
        cleaned = clean_row(raw, row_num)
        # Skip duplicate Contact_IDs
        cid = cleaned["Contact_ID"]
        if cid in seen_ids:
            skipped += 1
            continue
        seen_ids.add(cid)
        contacts_rows.append(row_to_sheet_values(cleaned))
        processed += 1

    print(f"  ✓ Processed: {processed} contacts")
    print(f"  ⏭ Skipped (empty/duplicate): {skipped}")

    # ── Users sheet ───────────────────────────────────────────────────────────
    print(f"\n👤 Creating admin user: {args.admin_email}")
    try:
        admin_row = create_admin_user(args.admin_password, args.admin_name, args.admin_email)
    except ImportError:
        sys.exit("❌ passlib not installed. Run: pip install passlib[bcrypt]")

    users_rows = [USERS_HEADERS, admin_row]

    # ── Audit log ─────────────────────────────────────────────────────────────
    audit_rows = [AUDIT_HEADERS]

    if args.dry_run:
        print("\n🔍 DRY RUN — no data written to Google Sheets")
        print(f"  Would write {len(contacts_rows) - 1} contacts to '{args.contacts_sheet}'")
        print(f"  Sample row: {contacts_rows[1] if len(contacts_rows) > 1 else 'N/A'}")
        print(f"  Would write admin user to '{args.users_sheet}'")
        return

    # ── Write to Sheets ───────────────────────────────────────────────────────
    print("\n🔗 Connecting to Google Sheets ...")
    service = build_sheets_service(args.credentials)

    print("\n📋 Setting up sheet tabs ...")
    for tab in [args.contacts_sheet, args.users_sheet, args.audit_sheet]:
        ensure_sheet_tab(service, args.spreadsheet_id, tab)

    print("\n✍️  Writing data ...")
    write_to_sheet(service, args.spreadsheet_id, args.contacts_sheet, contacts_rows)
    write_to_sheet(service, args.spreadsheet_id, args.users_sheet, users_rows)
    write_to_sheet(service, args.spreadsheet_id, args.audit_sheet, audit_rows)

    print(f"\n{'='*60}")
    print("✅ Migration complete!")
    print(f"{'='*60}")
    print(f"  Contacts migrated : {processed}")
    print(f"  Admin user email  : {args.admin_email}")
    print(f"  Admin password    : {args.admin_password}")
    print()
    print("⚠️  IMPORTANT: Change the admin password after first login!")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
